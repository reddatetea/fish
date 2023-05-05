import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Font,PatternFill,GradientFill,Side,Border,Alignment
import xlwings as xw
import easygui
import excelmessage
import os

def setPrintArea(fname,ws_name=None):        #打印区域设置
    wb = openpyxl.load_workbook(fname)
    if ws_name ==None:
        sheetnames = wb.sheetnames
        msg = '请点选要设置打印区域的excel工作表'
        if len(sheetnames) == 1:
            ws = wb.active
        else:
            print(msg)
            choice = easygui.buttonbox(msg, title='设置打印区域', choices=sheetnames)
            ws = wb['%s' % choice]
    else :
        ws = wb[ws_name]


    ws_name = ws.title

    max_row = ws.max_row      #最大行数
    #print(max_row)
    max_column = ws.max_column   #最大列数
    #print(max_column)

    max_column_letter = get_column_letter(max_column)    # 根据列的数字返回字母

    # 根据字母返回列的数字
    #print(column_index_from_string('D')) # 4

    max_cell = '$%s$%s'%(max_column_letter,max_row)   #max_row和mar_column对应的单元格
    print_area = '$A$1:{}'.format(max_cell)        #打印区域
    ws.print_area = print_area                     #设置打印区域

    wb.save(fname)
    return fname,ws_name


def firstRowSeting(fname,ws_name=None):        #首行设置、其它行设置
    wb = openpyxl.load_workbook(fname)
    if ws_name == None:
        sheetnames = wb.sheetnames
        msg = '请点选要设置首行的工作表'

        if len(sheetnames) == 1:
            ws = wb.active
        else:
            print(msg)
            choice = easygui.buttonbox(msg, title=msg, choices=sheetnames)
            ws = wb['%s' % choice]
    else :
        ws = wb[ws_name]


    ws_sheetname = ws.title

    max_row = ws.max_row

    #sheet_view.topLeftCell = 'A1'     #避坑
    ws.print_title_cols = 'A:B'  # the first two cols
    ws.print_title_rows = '1:1'  # the first row
    ws.freeze_panes = 'B2'
    # ws.row_dimensions[1].height = 20
    # ws.column_dimensions['B'].width = 20

    font = Font(name = '微软雅黑',size=12,bold=True,italic=True,color='000000')
    font1 = Font(name='微软雅黑', size=10, bold=False, italic=False, color='000000')
    pattern_fill = PatternFill(fill_type = 'solid',fgColor='40E0D0')  #绿宝石
    pattern_fill1 = PatternFill(fill_type='solid', fgColor='AFEEEE')  # 苍白的绿宝石
    alignment = Alignment(horizontal = 'center',vertical = 'center',text_rotation = 0,wrap_text = False)

    side = Side(style = None,color = '000000')        #边框 线型为thin,黑色
    border = Border(left = side,right = side,top = side,bottom = side)    #定义边框的左右上下

    for row in ws[1:ws.max_row]:
        for j in row:
            if j.row == 1 or j.row == ws.max_row :
                j.font = font
                j.border = border
                j.fill = pattern_fill
                j.alignment = alignment
            else :
                if j.row%2 == 0:
                    j.border = border
                    continue
                else :
                    j.border = border
                    j.fill = pattern_fill1
    wb.save(fname)

    app = xw.App(visible = False,add_book = False)
    wb = app.books.open(fname)
    ws = wb.sheets[ws_sheetname]
    ws.autofit()                #自动适应单元格
    wb.save(fname)
    wb.close()
    app.quit()
    return fname


def yemei(fname,ws_name=None,gongyingshang=None):
    wb = openpyxl.load_workbook(fname)
    if ws_name ==None:
        sheetnames = wb.sheetnames
        print('现在设置页眉页脚')
        if len(sheetnames) == 1:
            ws = wb.active
        else:

            msg = '请点选要设置页眉页脚的excel工作表'
            choice = easygui.buttonbox(msg,title = msg,choices = sheetnames)
            ws = wb['%s'%choice]
    else :
        ws = wb[ws_name]


    # ws.print_options.horizontalCentered = True
    # ws.print_options.verticalCentered = True
    if gongyingshang  == None :
        msg = '请选择是否输入标题'
        print(msg)
        titleYN = easygui.ccbox(msg, title='请选择Yes or No', choices=('yes', 'no'), image=None)
        if titleYN == True:
            title = input('请输入打印标题：\n')
        else:
            title = ws.title
    else :
        title = gongyingshang


    #冻结单元格
    ws.freeze_panes = 'B2'

    #打印标题
    ws.print_title_cols = 'A:B' # the first two cols   #左端标题列（即从左侧重复的列数）
    ws.print_title_rows = '1:1' # the first row     #顶端标题行（即从上端重复的行数）

    #页眉设置
    ws.oddHeader.center.text =  '%s'%title    #1/n      页眉中文字
    ws.oddHeader.center.size = 18                   # 页眉中字体大小
    # ws.oddHeader.center.font = '宋体,Italic'         # 页眉中字体,宋体、倾斜
    # ws.oddHeader.center.font = '宋体,Bold'         # 页眉中字体,宋体、加粗
    ws.oddHeader.center.font = '宋体,Bold Italic'     # 页眉中字体,宋体、加粗、倾斜

    ws.oddHeader.center.color = "000000"            # 页眉中字体颜色

    #页脚设置
    ws.oddFooter.center.text =  " &[Page] / &N"     #1/n      页脚中文字
    ws.oddFooter.center.size = 12                   # 页脚中字体大小
    ws.oddFooter.center.font = "Tahoma"             # 页脚中字体
    ws.oddFooter.center.color = "000000"            # 页脚中字体颜色

    ws.oddFooter.right.text =  "张文伟 &[Date]"     #页脚右 文字
    ws.oddFooter.right.size = 12                    #页脚右 字体大小
    ws.oddFooter.right.font = "书体坊米芾体"        #页脚右 字体
    ws.oddFooter.right.color = "000000"             #页脚右 字体颜色
    wb.save(fname)

def fastseting(fname,ws_name=None,gongyingshang=None):
    setPrintArea(fname, ws_name)
    firstRowSeting(fname, ws_name)
    yemei(fname,ws_name,gongyingshang)

def main():
    fname = excelmessage.wenjian()
    fname = excelmessage.excelMessage(fname)
    path,filename = os.path.split(fname)
    os.chdir(path)
    # gongyingshang = None
    # ws_name = None
    fname = setPrintArea(fname)
    fname = firstRowSeting(fname)
    yemei(fname)

if __name__ == '__main__':
    main()