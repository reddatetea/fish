'''
核对信华柔板的价格
'''
import roubanjisuan
import openpyxl
from openpyxl.utils import column_index_from_string
import tiqushuju
import qijiansplit
import os
import easygui
import zhixiangheji
import time

def roubanHedui(fname,ws_name,im_gongyingshang):
    wb = openpyxl.load_workbook(fname)
    ws = wb[ws_name]
    sheetnames = [ws_name]
    maxrow = ws.max_row
    print(maxrow)
    pinming_num = column_index_from_string('D')
    shuliang_num = column_index_from_string('F')
    hetong_danjia_num = column_index_from_string('S')
    hetong_jiner_num = column_index_from_string('T')
    ruku_jiner_num = column_index_from_string('H')
    danjia_duoji_num = column_index_from_string('U')
    jiner_duoji_num = column_index_from_string('V')


    for row in range(1,maxrow+1):
        if row == 1 :
            ws.cell(row,hetong_danjia_num,value='合同单价')
            ws.cell(row, hetong_jiner_num, value='合同金额')
            ws.cell(row, danjia_duoji_num, value='单价多计')
            ws.cell(row, jiner_duoji_num, value='金额多计')

        else :
            pinming = ws.cell(row,pinming_num).value
            print(pinming)
            chang, kuan = roubanjisuan.roubanchicun(pinming)
            hetong_danjia = roubanjisuan.jisuanDanjia(im_gongyingshang,chang,kuan,pinming)
            ws.cell(row,hetong_danjia_num,value = hetong_danjia)
            hetong_jiner = round(ws.cell(row,shuliang_num).value*hetong_danjia,2)
            ws.cell(row, hetong_jiner_num, value=hetong_jiner)
            danjia_duoji = ws.cell(row,shuliang_num+1).value - hetong_danjia
            ws.cell(row,danjia_duoji_num,value = danjia_duoji)
            jiner_duoji  = ws.cell(row,shuliang_num+2).value - hetong_jiner
            ws.cell(row,jiner_duoji_num,value = jiner_duoji)


    wb.save(fname)
    return fname,sheetnames

def main():
    #fname = easygui.fileopenbox(msg='请选择原材料实时流水账', title='实时流水账')
    fname = r'F:\a00nutstore\006\zw\原材料实时流水账\原材料实时流水账.xlsx'
    path, filename = os.path.split(fname)
    gongyingshangs = ['信华','辉盈']
    for im_gongyingshang in  gongyingshangs:
        # for im_gongyingshang in  gongyingshangs:
        im_qijian0 = easygui.enterbox(msg='请输入要查询的期间,如果查询多个期间，中间用and连接', title='期间')
        im_qijians = qijiansplit.shijiqijian(im_qijian0)
        fname1,ws_name = tiqushuju.tiquShuju(fname, im_gongyingshang, im_qijians)
        fname1,sheetnames = roubanHedui(fname1,ws_name, im_gongyingshang)
        letters = 'FHSTV'
        zhixiangheji.sumheji(fname1, sheetnames, letters)
        path, filename = os.path.split(fname1)
        choice = easygui.ccbox(msg='是否打开要查询的文件"%s"' % filename, title='{}柔板价格核对'.format(im_gongyingshang), choices=('yes', 'no'))
        if choice == True:
            os.startfile(fname1)
        else:
            pass



if __name__ == '__main__' :
    main()









